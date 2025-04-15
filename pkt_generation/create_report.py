"""
📊 DPDK Kubernetes Experiment Report Generator
=============================================

This script is a comprehensive reporting tool for experiments conducted using
`packet_generator.py`, a DPDK benchmarking suite deployed in Kubernetes.

🔧 Purpose
----------
It processes experiment data collected from TX and RX pods—stored in `.npz`, `.csv`, and `.log` files—
and generates  reports in Markdown, LaTeX, or Excel formats. These reports focus e:

- TX vs RX packet convergence
- Byte-level loss and packet loss statistics
- Errors during packet reception
- Multi-pod and multi-node topology layouts
- Aggregated statistics across pod pairs

📁 Directory Structure
----------------------
Each experiment directory is structured as follows:

results/
└── <exp_id>/
    ├── tx0-rx0/
    │   └── profile_<params>/
    │       ├── <exp_id>_tx0_tx_...npz     # TX pod stats
    │       ├── <exp_id>_rx0_rx_...npz     # RX pod stats
    │       ├── metadata.txt               # Run specification
    │       ├── tx0_port_stats.csv         # TX counters (DPDK)
    │       └── rx0_stats.log              # RX output log
    └── ...

📤 Output Formats
-----------------
The script supports 3 output formats:

1. `markdown` – Console-friendly format ideal for CLI analysis and GitHub reports.
2. `latex` – High-quality PDF-ready report including TikZ topology diagrams and summary tables.
3. `excel` – Tabular summary of test parameters and statistics for structured result analysis.

🛠️ Features
-----------
✔ Per-pod pair TX/RX PPS comparison plots
✔ Byte and packet loss statistics
✔ RX error tracking
✔ LaTeX TikZ diagram of pod-to-node topology
✔ Multi-experiment batch reporting
✔ Richly formatted Excel output for structured analysis

🚀 Usage
--------
Run the script with an experiment ID:

```bash
python report_generator.py e5f97e8d --format markdown
"""

import argparse
import os
import re
import subprocess
from typing import Optional, Tuple, Dict, List

import matplotlib.pyplot as plt
from scipy.interpolate import make_interp_spline
from datetime import datetime

from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont
import numpy as np


def load_npz_data(
        path: str
) -> Optional[np.lib.npyio.NpzFile]:
    """
     Load NPZ data from a file.

    This function attempts to load a `.npz` file (NumPy zipped archive)
    and return the data as a `NpzFile` object. If loading fails, it logs the error
    and returns `None`.

    :param path: path (str): Path to the `.npz` file.
    :return:  Optional[np.lib.npyio.NpzFile]: The loaded NPZ data if successful, otherwise None.
    """
    try:
        data = np.load(path)
        return data
    except Exception as e:
        print(f"❌ Error loading {path}: {e}")
        return None


def inspect_npz(path: str):
    """
    Inspect npz file
    :param path: path (str): Path to the `.npz` file.
    :return:
    """
    print(f"\n📂 Inspecting: {path}")
    data = load_npz_data(path)
    if data is None:
        return

    print(f"🔑 Keys: {list(data.keys())}")
    for key in data.files:
        print(f"\n📊 {key}:")
        print(data[key])


def format_bytes(
        num: float
) -> str:
    """Convert bytes into a human-readable format.
    :param num:
    :return:
    """
    for unit in ['B', 'KB', 'MB', 'GB', 'TB', 'PB']:
        if abs(num) < 1024:
            return f"{num:.2f} {unit}"
        num /= 1024
    return f"{num:.2f} PB"


def format_number_compact(num):
    """    Convert a large number into a compact human-readable string using
    Kpkts, Mpkts, or Gpkts units.
    :param num:
    :return:
    """
    for unit, divider in [("Gpkts", 1_000_000_000), ("Mpkt", 1_000_000), ("Kpkt", 1_000)]:
        if abs(num) >= divider:
            return f"{num / divider:.2f} {unit}"
    return f"{num} pkts"


def infer_tx_rx_files(
        exp_id: str,
        results_dir="results"
) -> Tuple[Optional[str], Optional[str]]:
    """Recursively find TX and RX .npz files for a given experiment ID.

    Expected filename format:
    <exp_id>_<pod>_<role>_tx_txcores_<TX_CORES>_rxcores_<RX_CORES>_spec_profile_<FLOWS>_flows_pkt_size_<PKT_SIZE>B_<RATE>_rate_s_<timestamp>.npz

    Example:
        44d72c38_tx0_tx_txcores_3_rxcores_5_spec_profile_100_flows_pkt_size_9000B_100_rate_s_20250327_082320.npz
        44d72c38_rx0_rx_txcores_3_rxcores_5_spec_profile_100_flows_pkt_size_9000B_100_rate_s_20250327_082320.npz


    :param exp_id: experiment ID, The 8-character experiment ID prefix in the filename.
    :param results_dir: Root directory containing nested experiment result files.
    :return:  (tx_file_path, rx_file_path) if found, otherwise (None, None)
    """
    tx_file = rx_file = None
    for dirpath, _, filenames in os.walk(results_dir):
        for fname in filenames:
            if exp_id in fname and fname.endswith(".npz"):
                full_path = os.path.join(dirpath, fname)
                parts = fname.split("_")
                if len(parts) >= 3:
                    role = parts[2].lower()
                    if role == "tx" and tx_file is None:
                        tx_file = full_path
                    elif role == "rx" and rx_file is None:
                        rx_file = full_path

    return tx_file, rx_file


def interpolate_to_match(
        source_array: np.ndarray,
        target_len: int
) -> np.ndarray:
    """Interpolates a source array to match the length of a target.
    :param source_array:
    :param target_len:
    :return:
    """
    source_x = np.linspace(0, 1, len(source_array))
    target_x = np.linspace(0, 1, target_len)
    return np.interp(target_x, source_x, source_array)


def smooth_line(
        x: np.ndarray,
        y: np.ndarray,
        points=300
):
    """Return smoothed x and y using B-spline interpolation.
    :param x: (np.ndarray): 1D array of x-coordinates.
    :param y: y (np.ndarray): 1D array of y-coordinates corresponding to `x`.
    :param points: points (int, optional): Number of points to generate in the smoothed output.
                Defaults to 300.
    :return: -Tuple[np.ndarray, np.ndarray]:  Tuple[np.ndarray, np.ndarray]: Smoothed x and y arrays of length `points`
    """
    if len(x) < 4 or len(y) < 4:
        return x, y

    x_smooth = np.linspace(x.min(), x.max(), points)
    spline = make_interp_spline(x, y, k=3)
    y_smooth = spline(x_smooth)
    return x_smooth, y_smooth


def extract_metadata(
        filename: str,
        full_path: Optional[str] = None
) -> Dict[str, str]:
    """  Extract metadata like TX/RX cores, flows, packet size, and rate from filename.

    Example filename:
    44d72c38_tx0_tx_txcores_3_rxcores_5_spec_profile_100_flows_pkt_size_9000B_100_rate_s_20250327_082320.npz

    :param filename:
    :param full_path:  Full path to the file. Used to infer pod_pair.
    :return:  Parsed metadata as key-value pairs.
    """
    parts = filename.rstrip(".npz").split('_')
    metadata = {}

    try:
        metadata["exp_id"] = parts[0]
        metadata["pod"] = parts[1]
        metadata["role"] = parts[2]

        for i, part in enumerate(parts):
            if part == "txcores":
                metadata["txcores"] = parts[i + 1]
            elif part == "rxcores":
                metadata["rxcores"] = parts[i + 1]
            elif part == "profile":
                metadata["num_flows"] = parts[i + 1]
            elif part == "pkt":
                size_part = parts[i + 2]  # e.g., "128B"
                if size_part.endswith("B"):
                    metadata["pkt_size"] = int(size_part[:-1])
            elif part == "rate":
                metadata["percent_rate"] = parts[i - 1]

        if full_path:
            path_parts = os.path.normpath(full_path).split(os.sep)
            for p in path_parts:
                if "-" in p and "tx" in p and "rx" in p:
                    metadata["pod_pair"] = p
                    break

            meta_path = os.path.join(os.path.dirname(full_path), "metadata.txt")
            if os.path.isfile(meta_path):
                with open(meta_path) as f:
                    for line in f:
                        line = line.strip()
                        if "=" in line and not line.startswith("#"):
                            k, v = line.split("=", 1)
                            metadata[k.strip()] = v.strip()

    except (IndexError, ValueError) as e:
        print(f"⚠️ Error parsing metadata from filename: {filename}")
        print(f"🔍 Details: {e}")

    return metadata


def plot_rx_tx_pps(
        tx_pps: np.ndarray,
        rx_pps: np.ndarray,
        metadata: dict,
        save_dir: str
) -> Optional[str]:
    """
        Plot TX and RX PPS over time, with smoothing and mean lines,
        and save the figure as a PNG file.

        This function interpolates and smooths the TX and RX packets-per-second (PPS)
        arrays, overlays their time series on a single plot, and highlights
        the mean values for quick comparison. It also generates a dynamic
        title and filename using provided metadata.

    :param tx_pps: Array of TX packets per second over time.
    :param rx_pps: Array of RX packets per second over time.
    :param metadata: Metadata dictionary with keys like
                                   'exp_id', 'pod_pair', 'txcores', 'rxcores',
                                   'pkt_size', 'percent_rate', etc.
    :param save_dir: Directory where the plot image will be saved. ( by default saved in each pod pair )
    :return:
    """
    if len(tx_pps) == 0 or len(rx_pps) == 0:
        print("❌ No valid data to plot.")
        return

    tx_pps_interp = interpolate_to_match(tx_pps, len(rx_pps))
    time = np.arange(len(rx_pps))

    tx_time_smooth, tx_pps_smooth = smooth_line(time, tx_pps_interp)
    rx_time_smooth, rx_pps_smooth = smooth_line(time, rx_pps)

    mean_rx = np.mean(rx_pps[rx_pps > 1000])
    mean_tx = np.mean(tx_pps)

    plt.figure(figsize=(12, 6))
    plt.plot(tx_time_smooth, tx_pps_smooth, label="TX PPS (Smoothed)", color="tab:orange")
    plt.plot(rx_time_smooth, rx_pps_smooth, label="RX PPS (Smoothed)", color="tab:blue")

    plt.axhline(mean_tx, linestyle="--", color="tab:orange", label=f"Mean TX PPS ({int(mean_tx):,})")
    plt.axhline(mean_rx, linestyle=":", color="tab:blue", label=f"Mean RX PPS ({int(mean_rx):,})")

    title_parts = [
        f"TX vs RX PPS Observation Over Time | Pod Pair: {metadata.get('pod_pair', 'unknown')} "
        f"| TX Core: {metadata.get('txcores', 'N/A')} | RX Core: {metadata.get('rxcores', 'N/A')}",
        f"Packet Size: {metadata.get('pkt_size', 'N/A')}B | Rate: {metadata.get('percent_rate', 'N/A')}% | "
        f"Flows: {metadata.get('num_flows', 'N/A')}"
    ]

    plot_name = (
        f"{metadata.get('exp_id', 'unknown')}_"
        f"{metadata.get('pod_pair', 'unknownpair')}_"
        f"txcores_{metadata.get('txcores', 'X')}_rxcores_{metadata.get('rxcores', 'Y')}_"
        f"spec_profile_{metadata.get('num_flows', 'N')}_flows_"
        f"pkt_size_{metadata.get('pkt_size', 'X')}B_"
        f"{metadata.get('percent_rate', 'R')}_rate_pps.png"
    )

    save_path = os.path.join(save_dir, plot_name)
    plt.title("\n".join(title_parts))
    plt.xlabel("Time (seconds)")
    plt.ylabel("Packets Per Second (PPS)")
    plt.grid(True)
    plt.legend()
    plt.tight_layout()
    plt.savefig(save_path)
    print(f"📈 Plot saved to: {save_path}")
    plt.close()

    return save_path


def sort_experiments(experiments):
    """Sort experiments according to the metadata dictionary.
    :param experiments:
    :return:
    """
    def sort_key(item):
        _, tx_file = item
        metadata = extract_metadata(os.path.basename(tx_file), full_path=tx_file)
        return (
            int(metadata.get("num_flows", 0)),
            int(metadata.get("pkt_size", 0)),
            int(metadata.get("percent_rate", 0))
        )

    return sorted(experiments, key=sort_key)


def discover_experiments(root_dir="results"):
    """Return a list of (exp_id, tx_file_path) tuples for sorting and processing."""
    experiments = []
    for dirpath, _, filenames in os.walk(root_dir):
        for fname in filenames:
            if fname.endswith(".npz") and "_tx_" in fname:
                match = re.match(r"^([a-f0-9]{8})_", fname)
                if match:
                    exp_id = match.group(1)
                    full_path = os.path.join(dirpath, fname)
                    experiments.append((exp_id, full_path))
    return experiments


def discover_experiment_ids(
        root_dir="results"):
    """Walk the directory and collect experiment IDs based on .npz TX/RX file pairs."""
    exp_ids = set()
    for dirpath, _, filenames in os.walk(root_dir):
        for fname in filenames:
            if fname.endswith(".npz"):
                match = re.match(r"^([a-f0-9]{8})_", fname)
                if match:
                    exp_ids.add(match.group(1))
    return sorted(exp_ids)


def generate_latex_report(
        metadata: dict,
        plot_filename: str,
        tx_pps: np.ndarray,
        rx_pps: np.ndarray,
        tx_bytes: np.ndarray,
        rx_bytes: np.ndarray,
        tx_packets: int,
        rx_packets: int,
        rx_errors: Optional[np.ndarray] = None,
        report_title: str = "Experiment Report",
) -> str:
    """Generate latex report per each experiment.

    :param metadata: a dict with experiment metadata
    :param plot_filename:
    :param tx_pps:  np.ndarray  observer tx pps during a run ( sampled over time )
    :param rx_pps:  np.ndarray  observer rx pps during a run ( sampled over time )
    :param tx_bytes: np.ndarray tx bytes observer during a run ( sampled over time )
    :param rx_bytes: np.ndarray rx bytes observer during a run ( sampled over time )
    :param tx_packets: np.ndarray total transmit packets observer during a run ( sampled over time )
    :param rx_packets: np.ndarray total received packets pps observer during a run ( sampled over time )
    :param rx_errors: np.ndarray rx errors observer during a run ( sampled over time )
    :param report_title: Title for report
    :return:
    """
    latex = []

    def truncate_node_name(name: str) -> str:
        return name.split("-")[-1] if name else "node"

    node_a = truncate_node_name(metadata.get("tx_node", ""))
    node_b = truncate_node_name(metadata.get("rx_node", ""))
    node_label = "Single-Node" if node_a == node_b else "Multi-Node"

    full_title = f"{report_title} ({node_label})"

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    latex.append(r"\section*{%s - %s}" % (full_title, metadata.get("exp_id", "unknown")))
    latex.append(r"{\small")
    latex.append(r"\textbf{Tool:} dpdk-k8s-cycling-bench\\")
    latex.append(r"\textbf{Author:} Mustafa Bayramov\\")
    latex.append(r"\textbf{Generated on:} %s" % timestamp)
    latex.append(r"}")
    latex.append(r"""
        \begin{center}
        \begin{minipage}{0.45\textwidth}
        \begin{itemize}
          \item \textbf{Pod Pair:} \texttt{%s}
          \item \textbf{TX Cores:} %s
          \item \textbf{RX Cores:} %s
          \item \textbf{TX Node:} %s
          \item \textbf{RX Node:} %s
          \item \textbf{TX Core List:} %s
          \item \textbf{RX Core List:} %s
          \item \textbf{Packet Size:} %sB
          \item \textbf{Flows:} %s
          \item \textbf{Rate:} %s%%
        \end{itemize}
        \end{minipage}
        \hfill
        \begin{minipage}{0.5\textwidth}
        \begin{center}
        \begin{tikzpicture}[
          txpod/.style={draw, fill=blue!10, rounded corners, minimum width=2.5cm, minimum height=1.2cm},
          rxpod/.style={draw, fill=green!10, rounded corners, minimum width=2.5cm, minimum height=1.2cm},
          worker/.style={draw, dashed, inner sep=0.4cm, rounded corners},
          arrow/.style={->, thick}
        ]

        \node[worker, label=above:{Node %s}] (workerA) at (0,0) {
          \begin{tikzpicture}
            \node[txpod] (tx) at (0,0) {TX: \texttt{%s}};
          \end{tikzpicture}
        };

        \node[worker, label=above:{Node %s}] (workerB) at (7,0) {
          \begin{tikzpicture}
            \node[rxpod] (rx) at (0,0) {RX: \texttt{%s}};
          \end{tikzpicture}
        };

        \draw[arrow] (workerA.east) -- node[above]{Flows: %s UDP} (workerB.west);
        \end{tikzpicture}
        \end{center}
        \end{minipage}
        \end{center}
        """ % (
        metadata.get("pod_pair", "N/A"),
        metadata.get("tx_numa", "N/A"),
        metadata.get("rx_numa", "N/A"),
        node_a,
        node_b,
        metadata.get("txcores", "N/A"),
        metadata.get("rxcores", "N/A"),
        metadata.get("pkt_size", "N/A"),
        metadata.get("num_flows", "N/A"),
        metadata.get("percent_rate", "N/A"),
        node_a,
        metadata.get("pod", "tx0"),
        node_b,
        metadata.get("pod_pair", "rx0").split("-")[-1],
        metadata.get("num_flows", "N/A"),
    ))

    latex.append(r"\subsection*{TX vs RX PPS}")
    relative_path = os.path.relpath(plot_filename, start="results")
    latex.append(r"\includegraphics[width=\linewidth]{%s}" % relative_path)

    latex.append(r"\subsection*{Summary Statistics}")
    latex.append(r"\renewcommand{\arraystretch}{1.0}")
    latex.append(r"\small")
    latex.append(r"""
        \begin{center}
        \begin{tabular}{|l@{\hspace{9pt}}r|}
        \hline
        \textbf{Metric} & \textbf{Value} \\
        \hline
        \textbf{TX PPS (mean)} & \textbf{%s} \\
        \textbf{RX PPS (mean >1k)} & \textbf{%s} \\
        \hline
        TX Bytes Sent & %s \\
        RX Bytes Received & %s \\
        Byte Loss & %s \\
        Byte Loss \%% & %.2f \\
        \hline
        TX Packets Sent & %s \\
        RX Packets Received & %s \\
        Packet Loss & %s \\
        Packet Loss \%% & %.2f \\
        RX Errors & %s \\
        \hline
        \end{tabular}
        \end{center}
        """ % (
        int(np.mean(tx_pps)),
        int(np.mean(rx_pps[rx_pps > 1000])),
        int(np.max(tx_bytes)), int(np.max(rx_bytes)),
        int(np.max(tx_bytes) - np.max(rx_bytes)),
        100 * (np.max(tx_bytes) - np.max(rx_bytes)) / np.max(tx_bytes),
        tx_packets, rx_packets,
        tx_packets - rx_packets,
        100 * (tx_packets - rx_packets) / tx_packets,
        int(np.sum(rx_errors)) if rx_errors is not None else 0
    ))

    return "\n".join(latex)


def generate_summary_markdown(
        metadata: dict,
        plot_filename: str,
        tx_pps: np.ndarray,
        rx_pps: np.ndarray,
        tx_bytes: np.ndarray,
        rx_bytes: np.ndarray,
        tx_packets: int,
        rx_packets: int,
        rx_errors: Optional[np.ndarray] = None,
) -> str:
    """
    Generate a markdown-formatted report for a single experiment.

    🔎 TX keys:
          ▸ mbits_tx
          ▸ pkts_tx
          ▸ obytes
          ▸ opackets
          ▸ imissed
          ▸ oerrors
          ▸ rx_nombuf
          ▸ pkts_rx
          ▸ mbits_rx
          ▸ ipackets
          ▸ ibytes
          ▸ ierrors
          ▸ port_rx_nombuf -- port stats
          ▸ port_opackets -- port stats
          ▸ port_imissed -- port stats
          ▸ port_ierrors -- port stats
          ▸ port_ibytes -- port stats
          ▸ port_oerrors -- port stats
          ▸ port_obytes -- port stats
          ▸ port_ipackets -- port stats

        🔎 RX keys:
          ▸ rx_pps
          ▸ tx_pps
          ▸ rx_bytes
          ▸ tx_bytes
          ▸ rx_packets
          ▸ tx_packets
          ▸ rx_errors
          ▸ rx_missed
          ▸ rx_nombuf
            """

    rx_pps_cleaned = rx_pps[rx_pps > 1000]
    tx_total_bytes = int(np.max(tx_bytes))
    rx_total_bytes = int(np.max(rx_bytes))
    byte_loss = tx_total_bytes - rx_total_bytes
    byte_loss_pct = (byte_loss / tx_total_bytes) * 100 if tx_total_bytes > 0 else 0

    pkt_loss = tx_packets - rx_packets
    pkt_loss_pct = (pkt_loss / tx_packets) * 100 if tx_packets > 0 else 0
    rx_error_total = int(np.sum(rx_errors)) if rx_errors is not None else 0

    md = []

    md.append(f"# 📊 Experiment Report - {metadata.get('exp_id', 'unknown')}\n")

    md.append(f"**Pod Pair**: `{metadata.get('pod_pair', 'N/A')}`  ")
    md.append(f"**TX Cores**: {metadata.get('txcores', 'N/A')}  ")
    md.append(f"**RX Cores**: {metadata.get('rxcores', 'N/A')}  ")
    md.append(f"**Packet Size**: {metadata.get('pkt_size', 'N/A')}B  ")
    md.append(f"**Flows**: {metadata.get('num_flows', 'N/A')}  ")
    md.append(f"**Rate**: {metadata.get('percent_rate', 'N/A')}%\n")

    md.append("---\n")
    md.append("## 📈 TX vs RX Convergence PPS (Expectation RX = TX)\n")
    md.append(f"![pps_plot]({plot_filename})\n")

    md.append("---\n")
    md.append("## 📊 Summary\n")
    md.append("| Metric                | Value              |")
    md.append("|-----------------------|--------------------|")
    md.append(f"| TX PPS (mean)         | {np.mean(tx_pps):,.0f} |")
    md.append(f"| RX PPS (mean >1k)     | {np.mean(rx_pps_cleaned):,.0f} |")
    md.append(f"| TX Bytes Sent         | {tx_total_bytes:,} |")
    md.append(f"| RX Bytes Received     | {rx_total_bytes:,} |")
    md.append(f"| Byte Loss             | {byte_loss:,} |")
    md.append(f"| Byte Loss %           | {byte_loss_pct:.2f}% |")
    md.append(f"| TX Packets Sent       | {tx_packets:,} |")
    md.append(f"| RX Packets Received   | {rx_packets:,} |")
    md.append(f"| Packet Loss           | {pkt_loss:,} |")
    md.append(f"| Packet Loss %         | {pkt_loss_pct:.2f}% |")
    md.append(f"| RX Errors             | {rx_error_total} |")

    return "\n".join(md)


def validate_required_keys(
        data: dict, keys:
        list[str], label:
        str, file_path: str
) -> bool:
    for key in keys:
        if key not in data:
            print(f"❌ Missing '{key}' in {label} file: {file_path}")
            return False
        if data[key] is None or len(data[key]) == 0:
            print(f"❌ '{key}' in {label} file is empty: {file_path}")
            return False
    return True


def console(tx_data, rx_data):
    """

    :param tx_data:
    :param rx_data:
    :return:
    """
    tx_bytes = tx_data.get("port_obytes") or tx_data.get("tx_bytes")

    if tx_data is None or rx_data is None:
        return

    tx_pps = tx_data.get("pkts_tx")
    if tx_pps is None:
        print("⚠️ TX .npz is missing 'pkts_tx'")
        return

    rx_pps = rx_data.get("rx_pps")
    if rx_pps is None:
        print("⚠️ RX .npz is missing 'rx_pps'")
        return

    # we clean noise i.e. initial ramp of pps, otherwise it will skew mean
    rx_pps_cleaned = rx_pps[rx_pps > 1000]

    tx_bytes = tx_data.get("port_obytes")
    if tx_bytes is None:
        tx_bytes = tx_data.get("tx_bytes")

    rx_bytes = rx_data.get("rx_bytes")

    if tx_bytes is None or rx_bytes is None:
        print("❌ Missing TX or RX byte counters.")
        return

    # both testpmd and pktgen counter start from 0 and move toward max(64bit)
    tx_total_bytes = int(np.max(tx_bytes))
    rx_total_bytes = int(np.max(rx_bytes))

    byte_loss = tx_total_bytes - rx_total_bytes
    byte_loss_pct = (byte_loss / tx_total_bytes) * 100 if tx_total_bytes > 0 else 0

    rx_errors = rx_data.get("rx_errors")
    rx_error_total = int(np.sum(rx_errors)) if rx_errors is not None else 0

    tx_packets = np.max(tx_data["port_opackets"])
    rx_packets = np.max(rx_data["rx_packets"])
    pkt_loss = tx_packets - rx_packets
    pkt_loss_pct = (pkt_loss / tx_packets) * 100 if tx_packets > 0 else 0

    print("\n📊 Summary:")
    print(f"🔸 TX PPS:            mean={np.mean(tx_pps):,.0f}, min={np.min(tx_pps):,}, max={np.max(tx_pps):,}")
    print(
        f"🔸 RX PPS:            mean={np.mean(rx_pps_cleaned):,.0f}, min={np.min(rx_pps_cleaned):,}, "
        f"max={np.max(rx_pps_cleaned):,}")
    print(f"🔸 TX Bytes Sent:     {tx_total_bytes:,}")
    print(f"🔸 RX Bytes Received: {rx_total_bytes:,}")
    print(f"🔸 BYTE LOSS:         {byte_loss:,}")
    print(f"🔸 BYTE LOSS %:       {byte_loss_pct:.2f}%")
    print(f"🔸 RX Errors:         {rx_error_total}")

    print(f"🔸 TX Packets Sent:     {tx_packets:,}")
    print(f"🔸 RX Packets Received: {rx_packets:,}")
    print(f"🔸 PKT LOSS:            {pkt_loss:,}")
    print(f"🔸 PKT LOSS %:          {pkt_loss_pct:.2f}%")
    print(f"🔸 RX Errors:           {rx_error_total}")


def create_latex_cluster_diagram(
        pod_pairs,
        metadatas=None
):
    """
    Dynamically generates LaTeX TikZ code that exactly matches the static working example,
    using per-node parentheses in the `fit=` argument to avoid TikZ errors.
    """
    tx_lines = []
    rx_lines = []
    arrow_lines = []

    tx_ids = []
    rx_ids = []

    def truncate_node_name(name):
        return name.split("-")[-1] if name else "node"

    print("\n🔍 Debug pod_pairs + metadatas")
    for i, (tx, rx) in enumerate(pod_pairs):
        print(f"[{i}] TX={tx['name']}  →  RX={rx['name']}")
        print(f"     ↳ TX cores: {metadatas[i].get('tx_numa')}")
        print(f"     ↳ RX cores: {metadatas[i].get('rx_numa')}")

    for i, (tx, rx) in enumerate(pod_pairs):
        tx_name = tx["name"]
        rx_name = rx["name"]

        tx_node = truncate_node_name(tx["node"])
        rx_node = truncate_node_name(rx["node"])

        y_pos = -i * 2

        tx_id = f"tx_{tx_name}"
        rx_id = f"rx_{rx_name}"

        tx_lines.append(f"    \\node[pod] ({tx_id}) at (0,{y_pos}) {{\\texttt{{TX: {tx_name}}}}};")
        rx_lines.append(f"    \\node[pod] ({rx_id}) at (6,{y_pos}) {{\\texttt{{RX: {rx_name}}}}};")

        tx_core_label = " Cores: " + metadatas[i].get("tx_numa", "") if metadatas else ""
        rx_core_label = " Cores: " + metadatas[i].get("rx_numa", "") if metadatas else ""

        tx_ids.append(tx_id)
        rx_ids.append(rx_id)

        tx_lines.append(
            f"    \\node[pod] ({tx_id}) at (0,{y_pos}) {{\\begin{{tabular}}{{c}}\\texttt{{TX: {tx_name}}} \\\\ \\scriptsize{{{tx_core_label}}}\\end{{tabular}}}};"
        )
        rx_lines.append(
            f"    \\node[pod] ({rx_id}) at (6,{y_pos}) {{\\begin{{tabular}}{{c}}\\texttt{{RX: {rx_name}}} \\\\ \\scriptsize{{{rx_core_label}}}\\end{{tabular}}}};"
        )

        arrow_lines.append(
            f"    \\draw[arrow] ({tx_id}.east) to node[midway, above, sloped, font=\\scriptsize]{{Flow {metadatas[i]['num_flows']}}} ({rx_id}.west);"
        )

    tx_fit_parts = " ".join(f"({x})" for x in ["labelA"] + tx_ids)
    rx_fit_parts = " ".join(f"({x})" for x in ["labelB"] + rx_ids)

    tx_label = truncate_node_name(pod_pairs[0][0]["node"])
    rx_label = truncate_node_name(pod_pairs[0][1]["node"])

    tx_lines.append(
        f"    \\node[pod] ({tx_id}) at (0,{y_pos}) {{\\begin{{tabular}}{{c}}\\texttt{{TX: {tx_name}}} \\\\ \\scriptsize{{{tx_core_label}}}\\end{{tabular}}}};"
    )
    rx_lines.append(
        f"    \\node[pod] ({rx_id}) at (6,{y_pos}) {{\\begin{{tabular}}{{c}}\\texttt{{RX: {rx_name}}} \\\\ \\scriptsize{{{rx_core_label}}}\\end{{tabular}}}};"
    )

    return "\n".join([
        r"\subsection*{Cluster Topology Overview}",
        r"\vspace*{0pt}",
        r"\begin{center}",
        r"  \resizebox{\linewidth}{!}{%",
        r"    \begin{tikzpicture}[",
        r"      remember picture,",
        r"      pod/.style={draw, fill=blue!5, rounded corners, minimum width=2.2cm, minimum height=1.2cm, remember picture},",
        r"      workerbox/.style={draw, dashed, rounded corners, inner sep=0.5cm},",
        r"      arrow/.style={->, thick, >=latex}",
        r"    ]",
        f"    \\node (labelA) at (0,1.5) {{\\textbf{{Node {tx_label}}}}};",
        *tx_lines,
        f"    \\node[workerbox, fit={tx_fit_parts}] (box_Worker_A) {{}};",
        f"    \\node (labelB) at (6,1.5) {{\\textbf{{Node {rx_label}}}}};",
        *rx_lines,
        f"    \\node[workerbox, fit={rx_fit_parts}] (box_Worker_B) {{}};",
        *arrow_lines,
        r"    \end{tikzpicture}",
        r"  }",
        r"\end{center}",
        r"\vspace*{0pt}"
    ])


def pad_to_length(arr: np.ndarray, target_len: int) -> np.ndarray:
    """Pads the input array with zeros at the end to reach target_len."""
    pad_len = target_len - len(arr)
    if pad_len > 0:
        return np.pad(arr, (0, pad_len), mode='constant')
    return arr


def format_bytes(num):
    for unit in ['B', 'KB', 'MB', 'GB', 'TB', 'PB']:
        if abs(num) < 1024:
            return f"{num:.2f} {unit}"
        num /= 1024
    return f"{num:.2f} PB"


def format_number_compact(num):
    for unit, divider in [("Gpkts", 1_000_000_000), ("Mpkt", 1_000_000), ("Kpkt", 1_000)]:
        if abs(num) >= divider:
            return f"{num / divider:.2f} {unit}"
    return f"{num} pkts"


def generate_summary_excel(
    tx_pps_mean: float,
    rx_pps_mean: float,
    total_tx_bytes: int,
    total_rx_bytes: int,
    byte_loss: int,
    total_tx_packets: int,
    total_rx_packets: int,
    pkt_loss: int,
    pkt_loss_pct: float,
    metadatas: List[Dict[str, str]],
    output_file: str = "results/structured_test_results.xlsx"
) -> None:
    """
    Generate an Excel summary report for a DPDK experiment.

    :param metadatas:
    :param pkt_loss_pct:
    :param pkt_loss:
    :param byte_loss:
    :param tx_pps_mean:
    :param rx_pps_mean:
    :param total_tx_bytes:
    :param total_rx_bytes:
    :param total_tx_packets:
    :param total_rx_packets:
    :param output_file:
    :return:
    """

    tx_pps_mpps = round(tx_pps_mean / 1_000_000, 2)
    rx_pps_mpps = round(rx_pps_mean / 1_000_000, 2)

    observed_stats = [
        ("Total TX PPS", f"{int(tx_pps_mean):,} ({tx_pps_mpps} Mpps)"),
        ("Total RX PPS", f"{int(rx_pps_mean):,} ({rx_pps_mpps} Mpps)"),
        ("Total TX Bytes", f"{total_tx_bytes:,} ({format_bytes(total_tx_bytes)})"),
        ("Total RX Bytes", f"{total_rx_bytes:,} ({format_bytes(total_rx_bytes)})"),
        ("TX Packets", f"{total_tx_packets:,} ({format_number_compact(total_tx_packets)})"),
        ("RX Packets", f"{total_rx_packets:,} ({format_number_compact(total_rx_packets)})"),
        ("Packet Loss", f"{pkt_loss:,}"),
        ("Packet Loss %", f"{pkt_loss_pct:.2f}"),
        ("Byte Loss", f"{format_bytes(byte_loss)}"),
    ]

    observed_block = "\n".join([f"{k}: {v}" for k, v in observed_stats])

    node_set = set()
    all_pairs = []
    multi_core = False

    for i, m in enumerate(metadatas):
        pod_pair = m.get("pod_pair", f"tx{i}-rx{i}")
        tx_node = m.get("tx_node")
        rx_node = m.get("rx_node")
        tx_cores = m.get("tx_numa", "")
        rx_cores = m.get("rx_numa", "")

        node_set.add(tx_node)
        node_set.add(rx_node)
        all_pairs.append(f"Pair: {pod_pair}  | TX Cores: {tx_cores} | RX Cores: {rx_cores}")

        try:
            tx_core_count = len(tx_cores.split())
            rx_core_count = len(rx_cores.split())
            if tx_core_count > 2 or rx_core_count > 2:
                multi_core = True
        except:
            pass

    node_summary = "Single Node" if len(node_set) == 1 else "Multi-Node"
    pod_summary = "Single-Pod" if len(metadatas) == 1 else "Multi-Pod"
    core_summary = "Multi-Core" if multi_core else "Single-Core"
    pkt_size = metadatas[0].get("pkt_size", "N/A")
    flows = metadatas[0].get("num_flows", "N/A")

    title_line = f"{node_summary}, {pod_summary}, {core_summary}. {pkt_size} Byte , {flows} Flows"
    metadata_description = title_line + "\n\n" + "\n".join(all_pairs)

    df = pd.DataFrame([{
        "Test Description": metadata_description,
        "Observed Statistics": observed_block,
        "VM / TCA TKG": "",
        "SLA Target": "",
        "Reference TCA/TGK": "",
        "Conclusion": ""
    }])

    if not os.path.exists(output_file):
        df.to_excel(output_file, index=False)
    else:
        wb = load_workbook(output_file)
        ws = wb.active
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)
        wb.save(output_file)

    wb = load_workbook(output_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_col=2):
        cell = row[1]
        lines = str(cell.value).split("\n")
        rich_lines = []
        bold_font = InlineFont(b=True)
        normal_font = InlineFont()

        for line in lines:
            is_bold = line.startswith("Total TX PPS") or line.startswith("Total RX PPS")
            rich_lines.append(TextBlock(text=line + "\n", font=bold_font if is_bold else normal_font))

        rich_lines[-1].text = rich_lines[-1].text.rstrip()
        cell.value = CellRichText(rich_lines)
        cell.alignment = Alignment(wrapText=True)

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        cell = row[0]
        text = str(cell.value)
        if not text:
            continue

        lines = text.split("\n")
        rich_lines = []
        bold_font = InlineFont(b=True)
        normal_font = InlineFont()

        for i, line in enumerate(lines):
            line = line.strip()
            if not line:
                continue
            is_bold = i == 0 or "Node" in line or "Flows" in line
            rich_lines.append(TextBlock(text=line + "\n", font=bold_font if is_bold else normal_font))

        if rich_lines:
            rich_lines[-1].text = rich_lines[-1].text.rstrip()
            try:
                cell.value = CellRichText(rich_lines)
            except Exception as e:
                print(f"❌ Failed to apply rich text to column A cell: {e}")
        cell.alignment = Alignment(wrapText=True, vertical="center")

    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['C'].width = 20

    wb.save(output_file)
    print(f"✅ Excel report saved: {output_file}")


def create_report(
        exp_id: str,
        is_debug: bool = False,
        output_format: str = "markdown",
        report_title: str = "Experiment Report",
):
    """
    :param exp_id:
    :param is_debug:
    :param output_format:
    :param report_title:
    :return:
    """
    experiments = discover_experiments()
    matching_tx_files = [tx for eid, tx in experiments if eid == exp_id]

    def tx_sort_key(tx_file):
        meta = extract_metadata(os.path.basename(tx_file), full_path=tx_file)
        print(
            f"🧩 Sorting key for {os.path.basename(tx_file)} → size={meta.get('pkt_size')}, "
            f"flows={meta.get('num_flows')}, rate={meta.get('percent_rate')}")
        try:
            return (
                int(meta.get("pkt_size", 0)),
                int(meta.get("num_flows", 0)),
                int(meta.get("percent_rate", 0))
            )
        except Exception as e:
            print(f"⚠️ Failed to extract sorting key from metadata: {e}")
            return 0, 0, 0

    matching_tx_files = sorted(matching_tx_files, key=tx_sort_key)

    if not matching_tx_files:
        print(f"❌ No TX .npz files found for experiment ID: {exp_id}")
        return

    combined_reports = []
    pod_pairs = []

    # this aggregated data for summary table
    total_tx_bytes = 0
    total_rx_bytes = 0
    total_tx_packets = 0
    total_rx_packets = 0

    all_tx_pps = []
    all_rx_pps = []

    metadatas = []

    for tx_file in matching_tx_files:
        tx_filename = os.path.basename(tx_file)
        tx_dir = os.path.dirname(tx_file)

        tx_parts = tx_filename.split("_")
        if len(tx_parts) < 3:
            print(f"⚠️ Unexpected TX filename format: {tx_filename}")
            continue

        tx_pod = tx_parts[1]  # e.g., "tx0"
        if not tx_pod.startswith("tx"):
            print(f"⚠️ TX pod name does not start with 'tx': {tx_pod}")
            continue

        rx_pod = tx_pod.replace("tx", "rx", 1)
        rx_filename = tx_filename.replace(f"{tx_pod}_tx_", f"{rx_pod}_rx_")
        rx_file = os.path.join(tx_dir, rx_filename)

        if not os.path.isfile(tx_file) or not os.path.isfile(rx_file):
            print(f"⚠️ TX or RX file not found for pod pair: {tx_pod} - {rx_pod}")
            continue

        print(f"📁 TX file: {tx_file}")
        print(f"📁 RX file: {rx_file}")

        if is_debug:
            print("📦 RX ------------------ ------------------ ")
            inspect_npz(rx_file)
            print("📦 TX ------------------ ------------------ ")
            inspect_npz(tx_file)

        tx_data = load_npz_data(tx_file)
        rx_data = load_npz_data(rx_file)
        if not tx_data or not rx_data or len(tx_data.files) == 0 or len(rx_data.files) == 0:
            print("⚠️ One of the .npz files is missing or empty. Skipping.")
            continue

        required_tx_keys = ["pkts_tx", "port_opackets", "port_obytes"]
        required_rx_keys = ["rx_pps", "rx_packets", "rx_bytes"]

        if not validate_required_keys(tx_data, required_tx_keys, "TX", tx_file):
            print("⚠️ Skipping report due to invalid TX data.")
            continue
        if not validate_required_keys(rx_data, required_rx_keys, "RX", rx_file):
            print("⚠️ Skipping report due to invalid RX data.")
            continue

        metadata = extract_metadata(os.path.basename(tx_file), full_path=tx_file)

        plot_dir = os.path.dirname(tx_file)

        plot_path = plot_rx_tx_pps(tx_data["pkts_tx"], rx_data["rx_pps"], metadata, plot_dir)
        if plot_path is None:
            print("⚠️ Skipping report due to failed plot.")
            continue

        tx_packets = int(np.max(tx_data["port_opackets"]))
        rx_packets = int(np.max(rx_data["rx_packets"]))
        tx_bytes = tx_data["port_obytes"]
        rx_bytes = rx_data["rx_bytes"]
        rx_errors = rx_data.get("rx_errors")

        total_tx_bytes += int(np.max(tx_bytes))
        total_rx_bytes += int(np.max(rx_bytes))

        total_tx_packets += tx_packets
        total_rx_packets += rx_packets

        all_tx_pps.append(tx_data["pkts_tx"])
        all_rx_pps.append(rx_data["rx_pps"])

        pod_pairs.append({"name": tx_pod, "node": metadata.get("tx_node", "Worker A")})
        pod_pairs.append({"name": rx_pod, "node": metadata.get("rx_node", "Worker B")})

        metadatas.append(metadata)

        if output_format == "latex":
            report = generate_latex_report(
                metadata, plot_path, tx_data["pkts_tx"], rx_data["rx_pps"],
                tx_bytes, rx_bytes, tx_packets, rx_packets, rx_errors,
                report_title=report_title
            )
        else:
            report = generate_summary_markdown(
                metadata, plot_path, tx_data["pkts_tx"], rx_data["rx_pps"],
                tx_bytes, rx_bytes, tx_packets, rx_packets, rx_errors
            )

        combined_reports.append(report)

    if not combined_reports:
        print("⚠️ No valid reports generated.")
        return

    if (output_format in ["latex", "excel"]) and all_tx_pps and all_rx_pps:
        target_len = min(len(arr) for arr in all_tx_pps)
        tx_pps_clipped = [arr[:target_len] for arr in all_tx_pps]

        agg_tx_pps = np.sum(np.vstack(tx_pps_clipped), axis=0)
        agg_rx_pps = np.sum(np.vstack(all_rx_pps), axis=0)

        agg_metadata = {
            "pod_pair": "aggregate",
            "txcores": "-",
            "rxcores": "-",
            "pkt_size": metadata.get("pkt_size", "N/A"),
            "percent_rate": metadata.get("percent_rate", "N/A"),
            "num_flows": metadata.get("num_flows", "N/A"),
            "exp_id": metadata.get("exp_id", "unknown")
        }

        agg_plot_path = plot_rx_tx_pps(agg_tx_pps, agg_rx_pps, agg_metadata, plot_dir)

        agg_diagram = create_latex_cluster_diagram(
            [(pod_pairs[i], pod_pairs[i + 1]) for i in range(0, len(pod_pairs), 2)],
            metadatas
        )

        agg_summary = agg_diagram + r"""
\newpage
\section*{Aggregated Summary (All Pod Pairs)}

\begin{center}
\begin{tabular}{|l|r|}
\hline
Metric & Value \\
\hline
Total TX Bytes & %s \\
Total RX Bytes & %s \\
Byte Loss & %s \\
TX Packets & %s \\
RX Packets & %s \\
Packet Loss & %s \\
Packet Loss \%% & %.2f \\
Total TX PPS & %.0f \\
Total RX PPS & %.0f \\
\hline
\end{tabular}
\end{center}

\includegraphics[width=\linewidth]{%s}
""" % (
            f"{total_tx_bytes:,}",
            f"{total_rx_bytes:,}",
            f"{total_tx_bytes - total_rx_bytes:,}",
            f"{total_tx_packets:,}",
            f"{total_rx_packets:,}",
            f"{total_tx_packets - total_rx_packets:,}",
            100 * (total_tx_packets - total_rx_packets) / total_tx_packets if total_tx_packets > 0 else 0,
            np.mean(agg_tx_pps),
            np.mean(agg_rx_pps[agg_rx_pps > 1000]),
            os.path.relpath(agg_plot_path, start="results")
        )

        combined_reports.append(agg_summary)

    if output_format == "excel" and all_tx_pps and all_rx_pps:
        byte_loss = total_tx_bytes - total_rx_bytes
        pkt_loss = total_tx_packets - total_rx_packets
        pkt_loss_pct = (pkt_loss / total_tx_packets) * 100 if total_tx_packets > 0 else 0

        generate_summary_excel(
            np.mean(agg_tx_pps),
            np.mean(agg_rx_pps[agg_rx_pps > 1000]),
            total_tx_bytes,
            total_rx_bytes,
            byte_loss,
            total_tx_packets,
            total_rx_packets,
            pkt_loss,
            pkt_loss_pct,
            metadatas=metadatas
        )
        return

    if output_format == "latex":
        return "\n\n\\newpage\n\n".join(combined_reports)
    else:
        return "\n\n---\n\n".join(combined_reports)


def main(cmd):
    """
    :param cmd:
    :return:
    """
    output_ext = "md" if cmd.format == "markdown" else "tex"
    report_path = os.path.join("results", f"report.{output_ext}")

    if cmd.exp_id:
        report = create_report(
            exp_id=cmd.exp_id,
            is_debug=cmd.format == "latex" and cmd.pdf,
            output_format=cmd.format,
            report_title=cmd.title,
        )
        if report:
            with open(report_path, "w") as f:
                f.write(report)

            print(f"\n📝 Report written to: {report_path}")
            if cmd.format == "latex" and cmd.pdf:
                subprocess.run(["pdflatex", os.path.basename(report_path)], cwd="results")
                print("📄 PDF compiled.")
        else:
            print("⚠️ Report generation failed.")
    else:

        print("🔎 No experiment ID provided — scanning all...")

        discovered = discover_experiments()
        grouped = {}
        for exp_id, tx_file in discovered:
            grouped.setdefault(exp_id, []).append(tx_file)

        def sort_key(eid):
            all_tx_files = grouped[eid]
            best_file = min(
                all_tx_files,
                key=lambda tx_file_: (
                    int(extract_metadata(os.path.basename(tx_file_), full_path=tx_file_).get("pkt_size", 0)),
                    int(extract_metadata(os.path.basename(tx_file_), full_path=tx_file_).get("num_flows", 0)),
                    int(extract_metadata(os.path.basename(tx_file_), full_path=tx_file_).get("percent_rate", 0))
                )
            )
            metadata = extract_metadata(os.path.basename(best_file), full_path=best_file)
            print(f"🧩 Sorting key for exp_id={eid} based on {os.path.basename(best_file)}")
            return (
                int(metadata.get("pkt_size", 0)),
                int(metadata.get("num_flows", 0)),
                int(metadata.get("percent_rate", 0))
            )

        sorted_eids = sorted(grouped.keys(), key=sort_key)

        all_reports = []
        for i, exp_id in enumerate(sorted_eids, 1):
            print(f"\n=== [{i}/{len(sorted_eids)}] Processing experiment ID: {exp_id} ===")
            report = create_report(
                exp_id=exp_id,
                is_debug=cmd.format == "latex" and cmd.pdf,
                output_format=cmd.format,
                report_title=cmd.title,
            )
            if report:
                all_reports.append(report)

        if all_reports:
            if cmd.format == "latex":
                header = r"""
    \documentclass{article}
    \usepackage{graphicx}
    \usepackage{tikz}
    \usepackage{enumitem}
    \usetikzlibrary{fit}
    \usepackage{geometry}
    \usetikzlibrary{shapes}
    \usetikzlibrary{arrows,arrows.meta}
    \usetikzlibrary{decorations.text,decorations.pathreplacing,patterns,shadows.blur,fadings}
    \usetikzlibrary{calc,quotes,positioning,fpu,angles,fit,spy,chains}
    \geometry{margin=1in}
    \begin{document}
    \setlist[itemize]{nosep, leftmargin=1.2em}

    """.strip()
                print(header)
                footer = r"\end{document}"
                full_report = header + "\n\n\\newpage\n\n".join(all_reports) + "\n\n" + footer
            else:
                full_report = "\n\n---\n\n".join(all_reports)

            with open(report_path, "w") as f:
                f.write(full_report)

            print(f"\n📝 Combined report written to: {report_path}")
            if cmd.format == "latex" and cmd.pdf:
                subprocess.run(["pdflatex", os.path.basename(report_path)], cwd="results", check=True)
                print("📄 PDF compiled.")
        else:
            print("⚠️ No valid reports generated.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="DPDK k8s experiment reporter")
    parser.add_argument("exp_id", nargs="?", help="Experiment ID to process")
    parser.add_argument("--format", choices=["markdown", "latex", "excel"], default="markdown",
                        help="Output report format")
    parser.add_argument("--pdf", action="store_true", help="📄 Compile LaTeX report to PDF")
    parser.add_argument("--title", default="Experiment Report", help="Title for the report section")
    args = parser.parse_args()
    main(args)
